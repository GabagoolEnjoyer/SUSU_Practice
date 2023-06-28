import sys

import openpyxl


print("!!ПРЕДУПРЕЖДЕНИЕ!!\n При работе со словарём, слова должны начинаться с большой буквы!")
def read_excel_to_dict(filename):
    data_dict = {} #Словарь для хранения данных
    workbook = openpyxl.load_workbook(filename)  #открываем файл Excel
    sheet = workbook.active #выбираем активный лист

    #проходим по строкам в файле
    for row in sheet.iter_rows(min_row=2, values_only=True):
        word = row[0]
        definition = row[1]
        example = row[2]

        data_dict[word] = [definition, example]

    #закрываем файл Excel
    workbook.close()
    return data_dict


def clear_excel(filename):
    workbook = openpyxl.Workbook()#Создаем новую книгу Excel
    workbook.remove(workbook.active)#Удаляем активный лист
    workbook.create_sheet("Словарь")  # Создаем новый лист
    workbook.save(filename)#Сохраняем изменения

def save_dicts_to_excel(filename, dictionaries):
    workbook = openpyxl.load_workbook(filename)#Загружаем существующую книгу Excel
    sheet = workbook.active
    row = 2  # Начальная строка для записи данных
    for word, data in dictionaries.items():
        sheet.cell(row=row, column=1, value=word)
        sheet.cell(row=row, column=2, value=data[0])
        sheet.cell(row=row, column=3, value=data[1])
        row += 1
    workbook.save(filename)  # Сохраняем изменения

filename = "dictionary.xlsx" #указываем имя файла Excel
excel_data = read_excel_to_dict(filename) #вызываем функцию для чтения данных из Excel в словарь


while True:
    flag = int(input("Введите: 1 - Найти, 2 - добавить, 3 - удалить, 4 - редактировать, 5 - сохранить и выйти\n"))
    match flag:
        case 1:
            search = input("Введите слово, которое хотите найти: ")
            if search in excel_data:
                print(excel_data[search])
            else:
                print("Слово не найдено.")
        case 2:
            new_word = input("Введите новое слово: ")
            new_definition = input("Введите определение: ")
            new_example = input("Введите пример использования слова: ")
            if new_word in excel_data:
                print("Ошибка! Данное слово уже присутствует в словаре!")
            else:
                excel_data[new_word] = [new_definition, new_example]
        case 3:
            del_word = input("Введите слово, которое хотите удалить: ")
            if del_word in excel_data:
                del excel_data[del_word]
            else:
                print("Ошибка! Данного слова нет в словаре!")
        case 4:
            redact_word = input("Введите слово, которое хотите отредактировать: ")
            if redact_word in excel_data:
                flag_redact = int(input("Введите 0 если хотите отредактировать определение,"
                                        " 1 если хотите отредактировать пример: "))
                redacted = input("Введите данные для редакции: ")
                excel_data[redact_word][flag_redact] = redacted
            else:
                print("Ошибка! Данного слова нет в словаре!")
        case 5:
            clear_excel(filename)
            save_dicts_to_excel(filename, excel_data)
            sys.exit()